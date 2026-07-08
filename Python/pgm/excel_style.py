"""
Unified Excel styling for all Python-generated Excel files.
Chinese text uses 宋体, English text uses Times New Roman.
Header row gets a colored background with white bold text.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_CN = "\u5B8B\u4F53"
FONT_EN = "Times New Roman"
HEADER_BG = "4338CA"
HEADER_FG = "FFFFFF"
BORDER_COLOR = "808080"

_thin_side = Side(style="thin", color=BORDER_COLOR)
_cell_border = Border(top=_thin_side, bottom=_thin_side, left=_thin_side, right=_thin_side)

_header_font = Font(name=FONT_EN, size=11, bold=True, color=HEADER_FG)
_header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

_body_font = Font(name=FONT_EN, size=10)
_body_align = Alignment(vertical="center")


def style_worksheet(ws):
    """Apply unified styling to a single openpyxl Worksheet."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
        for cell in row:
            cell.border = _cell_border
            if row_idx == 1:
                cell.font = _header_font
                cell.fill = _header_fill
                cell.alignment = _header_align
            else:
                cell.font = _body_font
                cell.alignment = _body_align

    if ws.max_row and ws.max_row >= 1:
        ws.row_dimensions[1].height = 24

    for col_idx in range(1, (ws.max_column or 0) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 3, 8), 50)
        ws.column_dimensions[col_letter].width = width


def style_workbook(wb):
    """Apply unified styling to all sheets in an openpyxl Workbook."""
    for ws in wb.worksheets:
        style_worksheet(ws)


def style_excel_file(file_path):
    """
    Open an existing .xlsx file, apply unified styling to all sheets, and save.
    Call this after pandas to_excel() / ExcelWriter to add formatting.
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    style_workbook(wb)
    wb.save(file_path)
    wb.close()
