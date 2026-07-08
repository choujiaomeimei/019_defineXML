import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy

CDISC_FILE = r"c:\Project_Web\019_defineXML\CDISC官方数据包\CDISC-SDTM-V3.4-English.xlsx"
SPEC_FILE  = r"c:\Project_Web\019_defineXML\CDISC官方数据包\sdtm_spec_output.xlsx"

# ── 1. 读取 CDISC 标准，构建查找表 ─────────────────────────────────
# lookup[(domain, varname)] = {ct_format, role, core, cdisc_notes}
cdisc_lookup = {}

cdisc_wb = openpyxl.load_workbook(CDISC_FILE, read_only=True)

SKIP_SHEETS = {"TOC", "TS_code", "Sheet6"}

# 映射: 我们的domain -> CDISC标准中的sheet
def get_cdisc_sheet(domain):
    if domain in cdisc_wb.sheetnames:
        return domain
    if domain.startswith("SUPP") and "SUPP--" in cdisc_wb.sheetnames:
        return "SUPP--"
    if domain.startswith("QS") and "QS" in cdisc_wb.sheetnames:
        return "QS"
    return None

for sn in cdisc_wb.sheetnames:
    if sn in SKIP_SHEETS:
        continue
    ws = cdisc_wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue

    header = [str(c).strip().replace("\n", " ") if c else "" for c in rows[0]]

    idx_vn = next((i for i, h in enumerate(header) if "Variable" in h and "Name" in h), None)
    idx_ct = next((i for i, h in enumerate(header) if "Controlled" in h), None)
    idx_role = next((i for i, h in enumerate(header) if h == "Role"), None)
    idx_notes = next((i for i, h in enumerate(header) if "Notes" in h), None)
    idx_core = next((i for i, h in enumerate(header) if h == "Core"), None)

    if idx_vn is None:
        continue

    for row in rows[1:]:
        if not row or not row[idx_vn]:
            continue
        varname = str(row[idx_vn]).strip().replace("\n", " ")
        if not varname or varname.startswith("Variable"):
            continue

        ct_format = str(row[idx_ct]).strip() if idx_ct is not None and row[idx_ct] else ""
        role      = str(row[idx_role]).strip() if idx_role is not None and row[idx_role] else ""
        notes     = str(row[idx_notes]).strip() if idx_notes is not None and row[idx_notes] else ""
        core      = str(row[idx_core]).strip() if idx_core is not None and row[idx_core] else ""

        cdisc_lookup[(sn, varname)] = {
            "ct_format": ct_format,
            "role": role,
            "core": core,
            "notes": notes,
        }

cdisc_wb.close()
print(f"CDISC标准共加载 {len(cdisc_lookup)} 条变量定义")


# ── 2. 查找函数 ────────────────────────────────────────────────────
def find_cdisc_info(domain, varname):
    """按优先级查找: 精确域 > SUPP--/QS模板 > 通用变量(多个域共有的变量)"""
    cdisc_sheet = get_cdisc_sheet(domain)
    if cdisc_sheet and (cdisc_sheet, varname) in cdisc_lookup:
        return cdisc_lookup[(cdisc_sheet, varname)]

    # 通用前缀匹配: 如 QSPH 的 QSSEQ -> QS 域的 --SEQ 或 QSSEQ
    if domain.startswith("QS") and "QS" in [k[0] for k in cdisc_lookup]:
        # 尝试把 QSPH 前缀的变量映射到 QS 的通用模式
        if ("QS", varname) in cdisc_lookup:
            return cdisc_lookup[("QS", varname)]
        # 变量名可能在 QS 模板中用 --XXX 表示, 如 QSTEST -> QS 中的 --TEST 等
        # 实际 QS sheet 里变量名就是 QSXXX, 直接查
        base = varname
        if len(domain) > 2 and varname.startswith(domain[:2]):
            generic_var = varname  # 如 QSSEQ -> 在 QS sheet 里就是 QSSEQ
            if ("QS", generic_var) in cdisc_lookup:
                return cdisc_lookup[("QS", generic_var)]

    # 尝试在所有域中找到相同变量名(通用变量如 STUDYID, USUBJID 等)
    for key, val in cdisc_lookup.items():
        if key[1] == varname:
            return val

    return None


# ── 3. Origin / Source 推断规则 ──────────────────────────────────────
def infer_origin_source(domain, varname, role, var_type):
    """基于SDTM变量特征推断Origin和Source"""
    vn = varname.upper()

    # STUDYID
    if vn == "STUDYID":
        return "Assigned", "Sponsor Defined"

    # DOMAIN / RDOMAIN
    if vn in ("DOMAIN", "RDOMAIN"):
        return "Assigned", "Sponsor Defined"

    # USUBJID
    if vn == "USUBJID":
        return "Derived", "DM.USUBJID"

    # SUBJID
    if vn == "SUBJID":
        return "CRF", "CRF"

    # --SEQ (序列号)
    if vn.endswith("SEQ"):
        return "Derived", "Sponsor Defined"

    # SUPP-- 特定变量
    if domain.startswith("SUPP"):
        if vn == "IDVAR":
            return "Assigned", "Sponsor Defined"
        if vn == "IDVARVAL":
            return "Derived", "Parent Domain"
        if vn == "QNAM":
            return "Assigned", "Sponsor Defined"
        if vn == "QLABEL":
            return "Assigned", "Sponsor Defined"
        if vn == "QVAL":
            return "CRF", "CRF"
        if vn == "QORIG":
            return "Assigned", "Sponsor Defined"
        if vn == "QEVAL":
            return "Assigned", "Sponsor Defined"

    # RELREC 特定变量
    if domain == "RELREC":
        if vn in ("IDVAR", "RELTYPE", "RELID"):
            return "Assigned", "Sponsor Defined"
        if vn == "IDVARVAL":
            return "Derived", "Parent Domain"

    # Trial Design 域 (TA, TE, TI, TV, TS)
    if domain in ("TA", "TE", "TI", "TV", "TS", "TD", "TM"):
        return "Protocol", "Protocol"

    # 日期时间变量 (以DTC结尾)
    if vn.endswith("DTC") or vn.endswith("DTM"):
        if vn in ("RFSTDTC", "RFENDTC", "RFXSTDTC", "RFXENDTC", "RFPENDTC", "RFICDTC"):
            return "Derived", "Derivation"
        if vn in ("DTHDTC", "BRTHDTC"):
            return "CRF", "CRF"
        return "CRF", "CRF"

    # 持续时间 (以DUR结尾)
    if vn.endswith("DUR"):
        return "Derived", "Derivation"

    # DY 变量 (Study Day)
    if vn.endswith("DY"):
        return "Derived", "Derivation"

    # 标志变量 (以FL结尾)
    if vn.endswith("FL"):
        return "Derived", "Derivation"

    # DM 特定变量
    if domain == "DM":
        if vn in ("AGE", "AGEU"):
            return "Derived", "Derivation"
        if vn in ("SEX", "RACE", "ETHNIC", "COUNTRY"):
            return "CRF", "CRF"
        if vn in ("SITEID",):
            return "CRF", "CRF"
        if vn in ("ARMCD", "ARM"):
            return "Protocol", "IXRS/Protocol"
        if vn in ("ACTARMCD", "ACTARM"):
            return "Derived", "Derivation"
        if vn in ("ARMNRS", "ACTARMUD"):
            return "Assigned", "Sponsor Defined"
        if vn == "DTHFL":
            return "Derived", "Derivation"

    # 标准化变量 (以STRF/ENRF/STR/END结尾的非日期变量)
    if vn.endswith("STRF") or vn.endswith("ENRF"):
        return "CRF", "CRF"

    # 通用 GRPID, SPID, REFID
    if vn.endswith("GRPID") or vn.endswith("SPID") or vn.endswith("REFID"):
        return "Assigned", "Sponsor Defined"

    # 通用 STAT, REASND
    if vn.endswith("STAT") or vn.endswith("REASND"):
        return "CRF", "CRF"

    # EPOCH
    if vn == "EPOCH":
        return "Derived", "Derivation"

    # VISITNUM, VISIT, VISITDY
    if vn in ("VISITNUM", "VISIT", "VISITDY"):
        return "Protocol", "Protocol"

    # TAETORD, ELEMENT
    if vn in ("TAETORD", "ELEMENT"):
        return "Derived", "Derivation"

    # 按Role推断
    if role == "Identifier":
        return "CRF", "CRF"
    if role == "Topic":
        return "CRF", "CRF"
    if role == "Record Qualifier":
        return "CRF", "CRF"
    if role == "Result Qualifier":
        return "CRF", "CRF"
    if role == "Synonym Qualifier":
        return "CRF", "CRF"
    if role == "Variable Qualifier":
        return "CRF", "CRF"
    if role == "Timing":
        return "CRF", "CRF"
    if role == "Rule":
        return "Assigned", "Sponsor Defined"

    return "CRF", "CRF"


# ── 4. 推断 CDISC Submission Value ──────────────────────────────────
def infer_submission_value(domain, varname, ct_format):
    """DOMAIN 变量的固定值, 以及 Controlled Terms 中的固定值"""
    vn = varname.upper()

    # DOMAIN 变量: 提交值就是域名本身
    if vn == "DOMAIN":
        actual_domain = domain
        if domain.startswith("SUPP"):
            actual_domain = domain
        return actual_domain

    # RDOMAIN: 依赖上下文, 留空
    if vn == "RDOMAIN":
        return ""

    # 如果 Controlled Terms 是一个固定字面值 (不包含括号、不是格式)
    # 例如 "DM", "AE" 等直接字面值
    if ct_format and not ct_format.startswith("(") and ct_format not in (
        "ISO 8601", "ISO 21090 NullFlavor enumeration"
    ) and len(ct_format) <= 8 and ct_format.isalnum():
        return ct_format

    return ""


# ── 5. 处理 sdtm_spec_output.xlsx ──────────────────────────────────
spec_wb = openpyxl.load_workbook(SPEC_FILE)

# 列索引 (0-based, 但openpyxl cell是1-based)
COL_MAP = {
    "Domain": 1, "Seq": 2, "Variable Name": 3, "Variable Label": 4,
    "Type": 5, "Length": 6, "Controlled Terms or Format": 7,
    "CDISC Submission Value": 8, "Origin": 9, "Source": 10,
    "Role": 11, "Core": 12, "SUPP": 13, "QEVAL": 14, "Text": 15,
    "CRF Page": 16, "Comment": 17, "Method": 18,
}

fill_yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

stats = {"matched": 0, "unmatched": 0, "total": 0}

for sn in spec_wb.sheetnames:
    ws = spec_wb[sn]
    domain = sn
    print(f"\n处理 {domain} ({ws.max_row - 1} 个变量):")

    for row_idx in range(2, ws.max_row + 1):
        varname = ws.cell(row=row_idx, column=COL_MAP["Variable Name"]).value
        var_type = ws.cell(row=row_idx, column=COL_MAP["Type"]).value or ""
        if not varname:
            continue

        stats["total"] += 1
        varname = str(varname).strip()

        info = find_cdisc_info(domain, varname)

        if info:
            stats["matched"] += 1

            # Controlled Terms or Format
            if info["ct_format"]:
                ws.cell(row=row_idx, column=COL_MAP["Controlled Terms or Format"]).value = info["ct_format"]

            # Role
            if info["role"]:
                ws.cell(row=row_idx, column=COL_MAP["Role"]).value = info["role"]

            # Core
            if info["core"]:
                ws.cell(row=row_idx, column=COL_MAP["Core"]).value = info["core"]

            # CDISC Submission Value
            ct_val = info["ct_format"]
            sub_val = infer_submission_value(domain, varname, ct_val)
            if sub_val:
                ws.cell(row=row_idx, column=COL_MAP["CDISC Submission Value"]).value = sub_val

            # Origin & Source
            origin, source = infer_origin_source(domain, varname, info["role"], var_type)
            ws.cell(row=row_idx, column=COL_MAP["Origin"]).value = origin
            ws.cell(row=row_idx, column=COL_MAP["Source"]).value = source

            print(f"  [OK] {varname}: CT={info['ct_format']!r}, Role={info['role']}, Core={info['core']}, Origin={origin}")
        else:
            stats["unmatched"] += 1

            # 即使没匹配到CDISC标准, 也根据变量名推断 Origin/Source
            origin, source = infer_origin_source(domain, varname, "", var_type)
            ws.cell(row=row_idx, column=COL_MAP["Origin"]).value = origin
            ws.cell(row=row_idx, column=COL_MAP["Source"]).value = source

            # DOMAIN 变量总是有提交值
            if varname.upper() == "DOMAIN":
                ws.cell(row=row_idx, column=COL_MAP["CDISC Submission Value"]).value = domain

            print(f"  [--] {varname}: not found in CDISC V3.4, Origin={origin}")

spec_wb.save(SPEC_FILE)
print(f"\n{'='*60}")
print(f"处理完成!")
print(f"  总变量数: {stats['total']}")
print(f"  匹配CDISC标准: {stats['matched']}")
print(f"  未匹配: {stats['unmatched']}")
print(f"  输出文件: {SPEC_FILE}")
