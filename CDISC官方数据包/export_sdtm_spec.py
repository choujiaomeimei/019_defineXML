import os
import glob
import pandas as pd
import pyreadstat
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

XPT_DIR = r"c:\Project_Web\019_defineXML\CDISC官方数据包\xpt"
OUTPUT_FILE = r"c:\Project_Web\019_defineXML\CDISC官方数据包\sdtm_spec_output.xlsx"

COLUMNS = [
    "Domain", "Seq", "Variable Name", "Variable Label", "Type", "Length",
    "Controlled Terms or Format", "CDISC Submission Value", "Origin",
    "Source", "Role", "Core", "SUPP", "QEVAL", "Text", "CRF Page",
    "Comment", "Method"
]

TYPE_MAP = {"string": "Char", "double": "Num", "int": "Num"}

xpt_files = sorted(glob.glob(os.path.join(XPT_DIR, "*.xpt")))

# 排除合并文件
xpt_files = [f for f in xpt_files if os.path.basename(f).upper() != "XPT_OUTPUT.XPT"]

print(f"共找到 {len(xpt_files)} 个XPT数据集")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    for xpt_path in xpt_files:
        fname = os.path.splitext(os.path.basename(xpt_path))[0].upper()
        print(f"  处理: {fname}")

        try:
            _, meta = pyreadstat.read_xport(xpt_path)
        except Exception as e:
            print(f"    读取失败: {e}")
            continue

        domain = meta.table_name or fname
        rows = []

        for seq, var_name in enumerate(meta.column_names, start=1):
            label = meta.column_names_to_labels.get(var_name, "")
            raw_type = meta.readstat_variable_types.get(var_name, "")
            var_type = TYPE_MAP.get(raw_type, raw_type)
            length = meta.variable_storage_width.get(var_name, "")

            rows.append({
                "Domain": domain,
                "Seq": seq,
                "Variable Name": var_name,
                "Variable Label": label,
                "Type": var_type,
                "Length": length,
                "Controlled Terms or Format": "",
                "CDISC Submission Value": "",
                "Origin": "",
                "Source": "",
                "Role": "",
                "Core": "",
                "SUPP": "",
                "QEVAL": "",
                "Text": "",
                "CRF Page": "",
                "Comment": "",
                "Method": "",
            })

        df = pd.DataFrame(rows, columns=COLUMNS)
        sheet_name = fname[:31]  # Excel sheet名最长31字符
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# 美化格式
wb = load_workbook(OUTPUT_FILE)
header_font = Font(bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_align = Alignment(vertical="center", wrap_text=False)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

col_widths = {
    "A": 8,   # Domain
    "B": 5,   # Seq
    "C": 18,  # Variable Name
    "D": 40,  # Variable Label
    "E": 6,   # Type
    "F": 8,   # Length
    "G": 25,  # Controlled Terms or Format
    "H": 22,  # CDISC Submission Value
    "I": 12,  # Origin
    "J": 12,  # Source
    "K": 12,  # Role
    "L": 8,   # Core
    "M": 8,   # SUPP
    "N": 12,  # QEVAL
    "O": 20,  # Text
    "P": 10,  # CRF Page
    "Q": 20,  # Comment
    "R": 20,  # Method
}

for ws in wb.worksheets:
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = cell_align

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

wb.save(OUTPUT_FILE)
print(f"\n导出完成: {OUTPUT_FILE}")
print(f"共 {len(wb.sheetnames)} 个sheet: {', '.join(wb.sheetnames)}")
